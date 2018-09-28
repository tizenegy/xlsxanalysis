import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import time
from time import gmtime, strftime
import glob
import os

def checkUnsupportedFunctions(setOfFunctions):
    open_file = open("admin/unsupported_functions.txt", "r")
    stri= ""
    for line in open_file:
        stri+=line 
    unsupportedList = stri.split()    #split the string and convert it into list
    open_file.close()
    myset = set()
    for unsup in unsupportedList:
        for found in setOfFunctions:
            if unsup.upper() not in found.upper():
                continue
            else:
                print("Found unsupported function: "+unsup)
                myset.add(unsup)
    return myset

def writeResultsToExcel(kpis):
    wb = load_workbook(filename='analysis/log.xlsx', read_only=False)
    #wb = Workbook()
    ws = wb.worksheets[0]
    ctr = 1
    firstEmptyRow = ws.max_row +1
    for kpi in kpis:
        ws.cell(row=firstEmptyRow, column=ctr).value = str(kpi)
        ctr=ctr+1
    wb.save(filename = 'analysis/log.xlsx')
    wb.close()
    return

# measure execution time
start_time = time.time()

# load the xlsx from the same dir as the py
files = glob.glob('./*.xlsx')
excel = files[0]
print("Loading Excel: "+str(excel))
wb = load_workbook(filename=excel, read_only=False,data_only=False)

# variables
funcSet = set()
unsupSet = set()
sheetCounter = 0
cellNotEmptyCounter = 0
refCounter = 0
formulaCounter = 0
numberOfCells=0

print("Performing analysis")
print("Depending on the size of the file, this might take a few minutes.")

# iterate thru xlsx and identify data, functions, and references
for sheet in wb.worksheets:
    numberOfCells = numberOfCells + (sheet.max_column*sheet.max_row)
print("Number of cells: "+str(numberOfCells))
for sheet in wb.worksheets:
    sheetCounter = sheetCounter+1
    print("Getting Sheet...")
    for row in sheet.rows:
        for cell in row:
            content = cell.value
            if cell.value:
                cellNotEmptyCounter=cellNotEmptyCounter+1
                if str(content).startswith("="):
                    refCounter = refCounter+1
                    if "(" in content:
                        formulaCounter = formulaCounter+1
                        funcSet.add(content)
                        testList.append(content)

# calculate kpis
testList = list()
fileSize = int(os.path.getsize(excel)/1000)
cellsEmpty = numberOfCells-cellNotEmptyCounter
cellsPlainData = cellNotEmptyCounter-refCounter
cellsRefs = int(refCounter-formulaCounter)
cellsUniqueFunc = len(funcSet)
cellsRedundantFunc = formulaCounter-cellsUniqueFunc
if formulaCounter != 0:
    funcRedundancy = int(cellsRedundantFunc/formulaCounter*100)
else:
    funcRedundancy = 0

if cellsPlainData != 0:
    calcPerData = round(formulaCounter/cellsPlainData,2)
else:
    calcPerData = 0

if cellsPlainData != 0:
    refPerData = round(cellsRefs/cellsPlainData,2)
else:
    refPerData = 0

unsupSet = checkUnsupportedFunctions(funcSet)

currentTime = strftime("%Y-%m-%d %Hh%Mm%Ss", gmtime())
print("Analysis complete")

# write report
file = open("analysis/Analysis "+currentTime+".txt", "a")
file.truncate(0)
file.write("\n"+"------------------------------------------------")
file.write("\n"+"Analysis of:        "+str(excel))
file.write("\n"+"File size:          "+str(fileSize))
file.write("\n"+"Number of sheets:   "+str(sheetCounter))
file.write("\n"+"Cells total:        "+str(numberOfCells))
file.write("\n"+"Cells empty:        "+str(cellsEmpty))
file.write("\n"+"Plain data:         "+str(cellsPlainData))
file.write("\n"+"Functions:          "+str(formulaCounter))
file.write("\n"+"Unique functions:   "+str(cellsUniqueFunc))
file.write("\n"+"References:         "+str(cellsRefs))
file.write("\n"+"-> Calculated cells per data cell:    "+str(calcPerData))
file.write("\n"+"-> Referencing cells per data cell:   "+str(refPerData))
file.write("\n"+"-> Redundancy of functions:           "+str(funcRedundancy)+"%")
file.write("\n"+"------------------------------------------------"+"\n"+"\n")
file.write("Unsupported functions found in sheet:"+"\n"+"\n")
for item in unsupSet:
    file.write(item+"\n")
file.write("\n"+"Unique functions in alphabetical order:"+"\n"+"\n")

for item in sorted(testList):
#for item in sorted(funcSet):
    file.write(item+"\n")


file.close()

kpiList = []
kpiList.extend([excel,fileSize,sheetCounter,numberOfCells,cellsEmpty,cellsPlainData,formulaCounter,cellsUniqueFunc,cellsRefs,calcPerData,refPerData,funcRedundancy])

writeResultsToExcel(kpiList)

print("Execution time: %6.2f seconds" % (time.time() - start_time))

input("Press Enter to exit...")